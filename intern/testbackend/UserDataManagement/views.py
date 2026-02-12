from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import Faculty, FacultyMapping, Student, DepartmentAdminAssignment
from .serializers import FacultySerializer,StudentPatchSerializer, DepartmentAdminAssignmentSerializer
from Creation.permissions import IsCollegeAdmin
from AcademicSetup.permissions import IsCampusAdmin
from datetime import datetime 
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.contrib.auth import get_user_model
from Creation.models import School, Department, Degree
from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q, F
from rest_framework import status
from .serializers import (
    FacultySerializer, 
    StudentPatchSerializer, 
    DepartmentAdminAssignmentSerializer,
    UserRoleSerializer
)


'''
-------------------------------------------------------------------------------------------------------------------------------
                                                Faculty creation
-------------------------------------------------------------------------------------------------------------------------------
'''


User = get_user_model()

REQUIRED_HEADERS = [
    "employee_id",
    "faculty_name",
    "faculty_email",
    "faculty_mobile_no",
    "faculty_date_of_birth",
    "faculty_gender",
    "dept_code",
    "school_code",
]

ALLOWED_GENDERS = ["MALE", "FEMALE", "OTHER"]


#Faculty Individual upload

class FacultyViewSet(viewsets.ModelViewSet):
    queryset = Faculty.objects.all()
    serializer_class = FacultySerializer
    lookup_field = "employee_id"   # 🔑 IMPORTANT

    def create(self, request, *args, **kwargs):
        mappings = request.data.pop("mappings", [])

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        faculty = serializer.save()



        return Response(
            FacultySerializer(faculty).data,
            status=status.HTTP_201_CREATED
        )

    def partial_update(self, request, *args, **kwargs):
        faculty = get_object_or_404(
            Faculty,
            employee_id=kwargs["employee_id"]
        )

        mappings = request.data.pop("mappings", None)

        serializer = self.get_serializer(
            faculty,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        faculty = serializer.save()

        if mappings is not None:
            FacultyMapping.objects.filter(faculty=faculty).delete()
            self._save_mappings(faculty, mappings)

        return Response(
            FacultySerializer(faculty).data,
            status=status.HTTP_200_OK
        )

    # 🔁 shared helper (clean + DRY)
    def _save_mappings(self, faculty, mappings):
        for mapping in mappings:
            school_code = mapping.get("school_code")
            dept_code = mapping.get("dept_code")

            school = School.objects.filter(
                school_code__iexact=school_code
            ).first()
            if not school:
                raise ValueError(f"Invalid school_code '{school_code}'")

            department = Department.objects.filter(
                dept_code__iexact=dept_code,
                degree__school=school
            ).first()
            if not department:
                raise ValueError(
                    f"Invalid dept_code '{dept_code}' for school '{school_code}'"
                )

            FacultyMapping.objects.create(
                faculty=faculty,
                school=school,
                department=department
            )

class FacultyMappingOptionsView(APIView):
    permission_classes = [IsCampusAdmin]

    def get(self, request):
        schools = School.objects.all().prefetch_related('degrees__departments')
        options = []

        for school in schools:
            options.append({
                "label": f"{school.school_name} (Full School)",
                "school_id": str(school.id),
                "department_id": None
            })

            for degree in school.degrees.all():
                for dept in degree.departments.all():
                    options.append({
                        "label": f"{school.school_name} - {dept.dept_name}",
                        "school_id": str(school.id),
                        "department_id": str(dept.id)
                    })

        return Response(options)


User = get_user_model()

REQUIRED_HEADERS = [
    "employee_id",
    "faculty_name",
    "faculty_email",
    "faculty_mobile_no",
    "faculty_date_of_birth",
    "faculty_gender",
    "dept_code",
    "school_code",
]

ALLOWED_GENDERS = ["MALE", "FEMALE", "OTHER"]


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()

    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except Exception:
            pass

    raise ValueError(
        f"Invalid date format '{value}'. Expected DD/MM/YYYY or YYYY-MM-DD"
    )


class FacultyBulkUploadAPIView(APIView):
    permission_classes = [IsCampusAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        excel_file = request.FILES.get("file")

        if not excel_file:
            return Response(
                {"error": "Excel file is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1] if cell.value]
        missing_headers = [h for h in REQUIRED_HEADERS if h not in headers]

        if missing_headers:
            return Response(
                {
                    "error": "Missing required columns",
                    "missing_columns": missing_headers
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        header_index = {h: headers.index(h) for h in REQUIRED_HEADERS}

        summary = {
            "total_rows": 0,
            "created_faculty": 0,
            "skipped_rows": 0,
        }

        row_errors = []

        for row_no, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if not any(cell.value for cell in row):
                continue

            summary["total_rows"] += 1

            data = {
                h: row[header_index[h]].value
                for h in REQUIRED_HEADERS
            }

            # ❗ STRICT REQUIRED CHECK
            missing_fields = [
                field for field in REQUIRED_HEADERS
                if not data.get(field)
            ]

            if missing_fields:
                summary["skipped_rows"] += 1
                row_errors.append({
                    "row": row_no,
                    "errors": [
                        "All columns must be filled: " +
                        ", ".join(REQUIRED_HEADERS)
                    ]
                })
                continue

            errors = []

            # ---------- GENDER ----------
            gender = str(data["faculty_gender"]).upper()
            if gender not in ALLOWED_GENDERS:
                errors.append(
                    f"Invalid gender '{data['faculty_gender']}'. "
                    f"Allowed: {ALLOWED_GENDERS}"
                )

            # ---------- DATE ----------
            try:
                dob = parse_excel_date(data["faculty_date_of_birth"])
            except ValueError as e:
                errors.append(str(e))
                dob = None

            # ---------- SCHOOL ----------
            school = School.objects.filter(
                school_code__iexact=str(data["school_code"]).strip()
            ).first()
            if not school:
                errors.append(
                    f"Invalid school_code '{data['school_code']}'"
                )

            # ---------- DEPARTMENT ----------
            department = None
            if school:
                department = Department.objects.filter(
                    dept_code__iexact=str(data["dept_code"]).strip(),
                    degree__school=school
                ).first()

                if not department:
                    errors.append(
                        f"Invalid dept_code '{data['dept_code']}' "
                        f"for school '{data['school_code']}'"
                    )

            # ---------- DUPLICATE ----------
            if Faculty.objects.filter(
                employee_id=data["employee_id"]
            ).exists():
                errors.append(
                    "Faculty with this employee_id already exists"
                )

            if errors:
                summary["skipped_rows"] += 1
                row_errors.append({
                    "row": row_no,
                    "errors": errors
                })
                continue

            # ---------- SAVE ----------

            user, user_created = User.objects.get_or_create(
                        username=data["employee_id"],
                        defaults={"email": data["faculty_email"]}
                    )

                    # Set password and role for newly created users
            if user_created:
                        user.set_password(data["employee_id"])  # password = employee_id
                        user.role = "FACULTY"
                        user.save()


            faculty, created = Faculty.objects.get_or_create(
                employee_id=data["employee_id"],
                defaults={
                    "user": user,
                    "faculty_name": data["faculty_name"],
                    "faculty_email": data["faculty_email"],
                    "faculty_mobile_no": data["faculty_mobile_no"],
                    "faculty_date_of_birth": dob,
                    "faculty_gender": gender,
                }
            )

            FacultyMapping.objects.get_or_create(
                faculty=faculty,
                school=school,
                department=department
            )

            if created:
                summary["created_faculty"] += 1

        return Response(
            {
                "status": "PARTIAL_SUCCESS" if row_errors else "SUCCESS",
                "summary": summary,
                "row_errors": row_errors
            },
            status=status.HTTP_200_OK
        )

"""
----------------------------------------------------------------------------------------------------------------------------
                                    Template Download for faculty
----------------------------------------------------------------------------------------------------------------------------
"""



class FacultyTemplateDownloadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Faculty Template"

        headers = [
            "employee_id",
            "faculty_name",
            "faculty_email",
            "faculty_mobile_no",
            "faculty_date_of_birth",
            "faculty_gender",
            "dept_code",
            "school_code",
        ]

        # ---- HEADER ROW ----
        ws.append(headers)

        bold_font = Font(bold=True)

        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index)
            cell.font = bold_font
            ws.column_dimensions[get_column_letter(col_index)].width = 25

        # ---- SAMPLE ROW (OPTIONAL BUT RECOMMENDED) ----
        ws.append([
            "MR1001",
            "Mr John Doe",
            "john.doe@example.com",
            "9876543210",
            "18/08/1999",
            "MALE",
            "CSE",
            "SOE",
        ])

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="faculty_upload_template.xlsx"'
        )

        wb.save(response)
        return response



"""
----------------------------------------------------------------------------------------------------------------------------------
                                                Filter for Faculty based on school_code and dept_code
----------------------------------------------------------------------------------------------------------------------------------
"""

class FacultyFilterAPIView(APIView):
    permission_classes = [IsCampusAdmin]

    def get(self, request):
        school_code = request.query_params.get("school_code")
        dept_code = request.query_params.get("dept_code")

        if not school_code and not dept_code:
            return Response(
                {
                    "error": "At least one filter is required: school_code or dept_code"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        mappings = FacultyMapping.objects.select_related(
            "faculty", "school", "department"
        )

        if school_code:
            mappings = mappings.filter(
                school__school_code__iexact=school_code
            )

        if dept_code:
            mappings = mappings.filter(
                department__dept_code__iexact=dept_code
            )

        faculty_qs = (
            Faculty.objects.filter(
                id__in=mappings.values_list("faculty_id", flat=True)
            )
            .distinct()
            .order_by("faculty_name")
        )

        serializer = FacultySerializer(faculty_qs, many=True)

        return Response(
            {
                "count": faculty_qs.count(),
                "results": serializer.data
            },
            status=status.HTTP_200_OK
        )




'''
--------------------------------------------------------------------------------------------------------------------------------
                                                Student creation
--------------------------------------------------------------------------------------------------------------------------------
'''


# ------------------------------------------
# STUDENTS
# ------------------------------------------

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
User = get_user_model()
from django.db import transaction, IntegrityError
from rest_framework import serializers
from .models import Semester,Regulation
from Creation.permissions import IsCollegeAdmin
from .models import Student
from .serializers import (
    StudentExcelUploadSerializer,
    StudentPatchSerializer,
    StudentCreateSerializer,
)


# ======================================================
# BULK POST – EXCEL UPLOAD
# ======================================================


class StudentExcelUploadAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCollegeAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get("file")

        if not file_obj:
            return Response(
                {"error": "Excel file is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = StudentExcelUploadSerializer(data={"file": file_obj})

        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = serializer.save()

            if result.get("errors"):
                return Response(
                    {
                        "status": "PARTIAL_SUCCESS",
                        "created": result.get("created", 0),
                        "skipped": result.get("skipped", []),
                        "errors": result.get("errors", []),
                    },
                    status=status.HTTP_200_OK,
                )

            return Response(
                {
                    "status": "SUCCESS",
                    "created": result.get("created", 0),
                    "skipped": result.get("skipped", []),
                    "errors": [],
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {
                    "status": "FAILED",
                    "error": f"Unexpected error during processing: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



# ======================================================
# BULK GET – ALL STUDENTS
# ======================================================
class StudentListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request):
        students = Student.objects.select_related(
            "department", "degree", "regulation", "semester"
        ).all()

        data = [
            {
                "roll_no": s.roll_no,
                "student_name": s.student_name,
                "student_email": s.student_email,
                "student_gender": s.student_gender,
                "student_date_of_birth": s.student_date_of_birth,
                "student_phone_number": s.student_phone_number,
                "department": s.department.dept_code,
                "regulation": s.regulation.regulation_code,
                "semester": s.semester.sem_number,
                "section": s.section,
                "is_active": s.is_active,
            }
            for s in students
        ]

        return Response(
            {
                "count": len(data),
                "results": data,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        serializer = StudentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {"message": "Student created successfully"},
            status=status.HTTP_201_CREATED,
        )


# ======================================================
# GET + PATCH – INDIVIDUAL STUDENT
# ======================================================
class StudentDetailAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request, roll_no):
        student = Student.objects.filter(roll_no=roll_no).first()
        if not student:
            return Response(
                {"error": "Student not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "roll_no": student.roll_no,
                "student_name": student.student_name,
                "student_email": student.student_email,
                "student_gender": student.student_gender,
                "student_date_of_birth": student.student_date_of_birth,
                "student_phone_number": student.student_phone_number,
                "parent_name": student.parent_name,
                "parent_phone_number": student.parent_phone_number,
                "batch": student.batch,
                "section": student.section,
                "department": student.department.dept_code,
                "regulation": student.regulation.regulation_code,
                "semester": student.semester.sem_number,
                "is_active": student.is_active,
            },
            status=status.HTTP_200_OK,
        )

    def patch(self, request, roll_no):
        student = Student.objects.filter(roll_no=roll_no).first()
        if not student:
            return Response(
                {"error": "Student not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = StudentPatchSerializer(
            instance=student,
            data=request.data,
            partial=True,
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(
            {"message": "Student updated successfully"},
            status=status.HTTP_200_OK,
        )



'''
--------------------------------------------------------------------------------------------------------------------------
                                            Download the student template
--------------------------------------------------------------------------------------------------------------------------
'''

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from django.http import HttpResponse
from openpyxl import Workbook

from Creation.permissions import IsCollegeAdmin


class StudentExcelTemplateDownloadAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Student Upload Template"

        headers = [
            "roll_no",
            "student_name",
            "student_email",
            "student_gender",
            "student_date_of_birth",
            "student_phone_number",
            "parent_name",
            "parent_phone_number",
            "regulation",
            "dept_code",
        ]

        # Write header row
        sheet.append(headers)

        # Optional: set column widths (nice UX)
        column_widths = [15, 20, 25, 15, 22, 22, 20, 22, 15, 15]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="student_upload_template.xlsx"'
        )

        workbook.save(response)
        return response
"""
--------------------------------------------------------------------------------------------------------------------------------
                                        Filtering students based on school code and dept code
--------------------------------------------------------------------------------------------------------------------------------
"""


class StudentFilterAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request):
        dept_code = request.query_params.get("dept_code")
        regulation_code = request.query_params.get("regulation")

        students = Student.objects.select_related(
            "department",
            "regulation",
            "semester",
        )

        # -----------------------------
        # FILTERING (DB FIELD BASED)
        # -----------------------------
        if dept_code:
            students = students.filter(
                department__dept_code__iexact=dept_code
            )

        if regulation_code:
            students = students.filter(
                regulation__regulation_code__iexact=regulation_code
            )

        degree_code = request.query_params.get("degree")
        if degree_code:
            students = students.filter(
                degree__degree_code__iexact=degree_code
            )

        section = request.query_params.get("section")
        if section:
            students = students.filter(
                section__iexact=section
            )

        # -----------------------------
        # RESPONSE FORMAT
        # -----------------------------
        data = [
            {
                "roll_no": s.roll_no,
                "student_name": s.student_name,
                "student_email": s.student_email,
                # 👇 response key can be anything
                "department": s.department.dept_code,
                "regulation": s.regulation.regulation_code,
                "semester": s.semester.sem_number,
                "section": s.section,
                "is_active": s.is_active,
            }
            for s in students
        ]

        return Response(
            {
                "count": len(data),
                "results": data,
            },
            status=status.HTTP_200_OK,
        )


# ==================================================================================
# DEPARTMENT ADMIN ASSIGNMENT VIEWS
# ==================================================================================
# These views handle the Department Admin assignment feature.
#
# Key Components:
# 1. DepartmentAdminAssignmentViewSet: CRUD operations for assignments
# 2. DegreesForSchoolView: Cascading filter - get degrees for a school
# 3. DepartmentsForDegreeView: Cascading filter - get departments for a degree
# 4. FacultySearchView: Search faculty by name or employee_id
#
# Cascading Selection Flow:
# Step 1: User selects School -> Frontend calls DegreesForSchoolView
# Step 2: User selects Degree -> Frontend calls DepartmentsForDegreeView
# Step 3: User selects Department(s) and Faculty
# Step 4: User clicks "Assign" -> Triggers MFA (handled in frontend)
# Step 5: After MFA success -> POST to DepartmentAdminAssignmentViewSet
# ==================================================================================

class DepartmentAdminAssignmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Department Admin assignments.
    
    Only Campus Admins can create, view, update, or delete assignments.
    When a new assignment is created, the faculty member's user role is
    automatically updated to DEPARTMENT_ADMIN (handled in model.save()).
    """
    queryset = DepartmentAdminAssignment.objects.all().select_related(
        'faculty', 'school', 'degree', 'department', 'assigned_by'
    )
    serializer_class = DepartmentAdminAssignmentSerializer
    permission_classes = [IsCampusAdmin]
    lookup_field = 'assignment_id'
    
    def perform_create(self, serializer):
        """
        Override perform_create to automatically set assigned_by to current user.
        
        This ensures we always know which campus admin made each assignment.
        """
        serializer.save(assigned_by=self.request.user)


class DegreesForSchoolView(APIView):
    """
    Cascading filter endpoint: Get all degrees for a selected school.
    
    Usage: GET /users/dept-admin/degrees-for-school/?school_id=<uuid>
    
    This is used in the frontend to populate the Degree dropdown after
    a School is selected, ensuring only relevant degrees are shown.
    """
    permission_classes = [IsCampusAdmin]
    
    def get(self, request):
        school_id = request.query_params.get('school_id')
        
        if not school_id:
            return Response(
                {"detail": "school_id query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all degrees for this school
        degrees = Degree.objects.filter(school_id=school_id).values(
            'degree_id', 'degree_name', 'degree_code'
        )
        
        return Response(list(degrees))


class DepartmentsForDegreeView(APIView):
    """
    Cascading filter endpoint: Get all departments for a selected degree.
    
    Usage: GET /users/dept-admin/departments-for-degree/?degree_id=<uuid>
    
    This is used in the frontend to populate the Department dropdown after
    a Degree is selected, ensuring only relevant departments are shown.
    """
    permission_classes = [IsCampusAdmin]
    
    def get(self, request):
        degree_id = request.query_params.get('degree_id')
        
        if not degree_id:
            return Response(
                {"detail": "degree_id query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all departments for this degree
        departments = Department.objects.filter(degree_id=degree_id).values(
            'dept_id', 'dept_name', 'dept_code'
        )
        
        return Response(list(departments))


class FacultySearchView(APIView):
    """
    Search endpoint: Find faculty by name or employee ID.
    
    Usage: GET /users/dept-admin/search-faculty/?q=<search_term>
    
    This powers the faculty search functionality in the assignment form.
    Supports partial matching on both full_name and employee_id.
    Only returns active faculty members.
    """
    permission_classes = [IsCampusAdmin]
    
    def get(self, request):
        search_query = request.query_params.get('q', '').strip()
        
        if len(search_query) < 2:
            return Response(
                {"detail": "Search query must be at least 2 characters"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Search by faculty_name or employee_id (case-insensitive)
        faculty = Faculty.objects.filter(
            is_active=True
        ).filter(
            Q(faculty_name__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        ).values(
            'id', 'faculty_name', 'employee_id', 'faculty_email'
        )[:20]  # Limit to 20 results
        
        return Response(list(faculty))

# ==================================================================================
# ROLES DASHBOARD VIEWS (FEATURE 4)
# ==================================================================================

class RolesSummaryView(APIView):
    """
    API for the Roles Dashboard Summary.
    Returns counts for Faculty, Campus Admins (CA), and Students.
    """
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self):
        counts = User.objects.aggregate(
            total_students=Count('id', filter=Q(role='STUDENT')),
            total_faculty=Count('id', filter=Q(role__in=['FACULTY', 'ACADEMIC_COORDINATOR', 'accedemic_coordinator'])),
            total_ca=Count('id', filter=Q(role__in=['COLLEGE_ADMIN', 'ACADEMIC_COORDINATOR', 'accedemic_coordinator']))
        )
        return Response(counts)

class RolesListView(APIView):
    """
    API for the Roles Dashboard Detail List.
    Supports filtering by role and academic hierarchy (for students).
    """
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request):
        role_filter = request.query_params.get('role')
        school_id = request.query_params.get('school_id')
        degree_id = request.query_params.get('degree_id')
        dept_id = request.query_params.get('department_id')
        batch = request.query_params.get('batch')
        search = request.query_params.get('search')

        queryset = User.objects.all().select_related(
            'faculty_profile', 
            'student_profile', 
            'student_profile__department', 
            'student_profile__degree', 
            'student_profile__degree__school'
        )

        # Role Filter
        if role_filter and role_filter != 'ALL':
            if role_filter == 'CA':
                queryset = queryset.filter(role__in=['COLLEGE_ADMIN', 'ACADEMIC_COORDINATOR'])
            else:
                queryset = queryset.filter(role=role_filter)

        # Academic Filters (Only for Students or if role filter is broad)
        is_student_view = role_filter == 'STUDENT'
        is_broad_view = not role_filter or role_filter == 'ALL'

        if is_student_view or is_broad_view:
            # We apply student-specific filters if they are provided
            if school_id:
                queryset = queryset.filter(Q(student_profile__degree__school_id=school_id) | ~Q(role='STUDENT'))
            if degree_id:
                queryset = queryset.filter(Q(student_profile__degree_id=degree_id) | ~Q(role='STUDENT'))
            if dept_id:
                queryset = queryset.filter(Q(student_profile__department_id=dept_id) | ~Q(role='STUDENT'))
            if batch:
                queryset = queryset.filter(Q(student_profile__batch=batch) | ~Q(role='STUDENT'))
            
            # If we are specifically filtering for students, ensure we only get students
            if is_student_view:
                queryset = queryset.filter(role='STUDENT')

        # Search
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(faculty_profile__faculty_name__icontains=search) |
                Q(student_profile__student_name__icontains=search)
            )

        serializer = UserRoleSerializer(queryset[:100], many=True) # Limit to 100 for performance
        return Response(serializer.data)


# ==================================================================================
# DASHBOARD STATS VIEW (FEATURE 5)
# ==================================================================================

class DashboardStatsView(APIView):
    """
    API for the College Admin Dashboard Statistics.
    Returns counts for:
    - Active Schools
    - Active Departments
    - Active Faculty
    - Active Students
    """
    permission_classes = [IsAuthenticated, IsCollegeAdmin]

    def get(self, request):
        stats = {
            "total_schools": School.objects.count(),
            "total_departments": Department.objects.count(),
            "total_faculty": Faculty.objects.filter(is_active=True).count(),
            "total_students": Student.objects.filter(is_active=True).count(),
        }
        return Response(stats)
