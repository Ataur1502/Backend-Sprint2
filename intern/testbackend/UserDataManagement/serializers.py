import csv
import io
import random
import string
from datetime import datetime
from openpyxl import load_workbook
from rest_framework import serializers
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate
from django.db import transaction
from .models import Faculty, FacultyMapping, Student,DepartmentAdminAssignment
from Creation.models import School, Department,Degree, Department, Regulation, Semester
from custom_auth.models import MFASession
from django.utils import timezone

User = get_user_model()

'''
---------------------------------------------------------------------------------------------------------------------------------
                                            Faculty creation
---------------------------------------------------------------------------------------------------------------------------------
'''



User = get_user_model()


# -------------------------------
# Mapping Serializers
# -------------------------------

class FacultyMappingCreateSerializer(serializers.Serializer):
    school_code = serializers.CharField()
    dept_code = serializers.CharField()


class FacultyMappingReadSerializer(serializers.ModelSerializer):
    school_code = serializers.CharField(
        source="school.school_code",
        read_only=True
    )
    dept_code = serializers.CharField(
        source="department.dept_code",
        read_only=True
    )

    class Meta:
        model = FacultyMapping
        fields = ["school_code", "dept_code"]


# -------------------------------
# Faculty Serializer
# -------------------------------

class FacultySerializer(serializers.ModelSerializer):
    #explicitly declare mappings (this was missing)
    mappings = FacultyMappingCreateSerializer(
        many=True,
        write_only=True,
        required=False
    )

    class Meta:
        model = Faculty
        fields = [
            "id",
            "employee_id",
            "faculty_name",
            "faculty_email",
            "faculty_mobile_no",
            "faculty_date_of_birth",
            "faculty_gender",
            "is_active",
            "created_at",
            "mappings",   
        ]
        read_only_fields = ["id", "created_at"]

    # -------------------------------
    # Validations
    # -------------------------------

    def validate_employee_id(self, value):
        qs = Faculty.objects.filter(employee_id=value)
        if self.instance:
            qs = qs.exclude(pk=self.instance.pk)

        if qs.exists():
            raise serializers.ValidationError(
                "A faculty with this employee_id already exists."
            )
        return value

    # -------------------------------
    # Create
    # -------------------------------

    @transaction.atomic
    def create(self, validated_data):
        mappings_data = validated_data.pop("mappings", [])
        employee_id = validated_data["employee_id"]
        email = validated_data.get("faculty_email")

        # ---------- USER ----------
        employee_id = validated_data["employee_id"]
        email = validated_data.get("faculty_email")
        user = User.objects.filter(username=employee_id).first()

        if user:
            # User already exists → FORCE role update
            user.email = email
            user.role = "FACULTY"          
            user.set_password(employee_id)
            user.is_active = True
            user.save()
        else:
            # New user → create as FACULTY
            user = User.objects.create(
                username=employee_id,
                email=email,
                role="FACULTY",             
                is_active=True,
            )
            user.set_password(employee_id)
            user.save()



        # ---------- FACULTY ----------
        faculty = Faculty.objects.create(
            user=user,
            **validated_data
        )

        # ---------- MAPPINGS ----------
        for mapping in mappings_data:
            school_code = mapping.get("school_code")
            dept_code = mapping.get("dept_code")

            school = School.objects.filter(
                school_code__iexact=school_code
            ).first()
            if not school:
                raise serializers.ValidationError(
                    {"school_code": f"Invalid school_code '{school_code}'"}
                )

            department = Department.objects.filter(
                dept_code__iexact=dept_code,
                degree__school=school
            ).first()
            if not department:
                raise serializers.ValidationError(
                    {
                        "dept_code": (
                            f"Invalid dept_code '{dept_code}' "
                            f"for school '{school_code}'"
                        )
                    }
                )

            FacultyMapping.objects.create(
                faculty=faculty,
                school=school,
                department=department
            )

        return faculty





'''
--------------------------------------------------------------------------------------------------------------------------------
                                                Student creation
--------------------------------------------------------------------------------------------------------------------------------
'''


"""
creates Students using .Xlsx file 
"""

from django.db import transaction
from rest_framework import serializers
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from .models import Student, Department, Regulation, Semester

User = get_user_model()

class StudentExcelUploadSerializer(serializers.Serializer):
    file = serializers.FileField()

    REQUIRED_COLUMNS = [
        "roll_no",
        "student_name",
        "student_email",
        "student_gender",
        "student_date_of_birth",
        "student_phone_number",
        "parent_name",
        "parent_phone_number",
        "regulation",
        "dept_code",
        "section",
    ]

    def save(self, **kwargs):
        workbook = load_workbook(self.validated_data["file"])
        sheet = workbook.active

        created = 0
        skipped = []
        errors = []

        for row_no, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):

            # -------------------------------------------------
            # 1️⃣ BASIC TEMPLATE CHECK
            # -------------------------------------------------
            if not row or len(row) < len(self.REQUIRED_COLUMNS):
                errors.append({
                    "row": row_no,
                    "error": "Column count mismatch with template",
                })
                continue

            # -------------------------------------------------
            # 2️⃣ COLLECT ALL MISSING COLUMN ERRORS (ONE LINE)
            # -------------------------------------------------
            missing_columns = []

            for idx, column_name in enumerate(self.REQUIRED_COLUMNS):
                if row[idx] in (None, "", " "):
                    missing_columns.append(f"{column_name} is required")

            if missing_columns:
                errors.append({
                    "row": row_no,
                    "error": ", ".join(missing_columns),
                })
                continue

            # -------------------------------------------------
            # 3️⃣ SAFE UNPACKING
            # -------------------------------------------------
            (
                roll_no,
                student_name,
                student_email,
                student_gender,
                student_date_of_birth,
                student_phone_number,
                parent_name,
                parent_phone_number,
                regulation_code,
                dept_code,
                section,
            ) = row[:11]


            # -------------------------------------------------
            # 4️⃣ DUPLICATE STUDENT CHECK
            # -------------------------------------------------
            if Student.objects.filter(roll_no=roll_no).exists():
                skipped.append({
                    "row": row_no,
                    "roll_no": roll_no,
                    "error": "Student already exists",
                })
                continue

            # -------------------------------------------------
            # 5️⃣ FOREIGN KEY VALIDATIONS
            # -------------------------------------------------
            department = Department.objects.filter(
                dept_code__iexact=dept_code
            ).first()
            if not department:
                errors.append({
                    "row": row_no,
                    "error": "Invalid department code",
                })
                continue

            regulation = Regulation.objects.filter(
                regulation_code__iexact=regulation_code
            ).first()
            if not regulation:
                errors.append({
                    "row": row_no,
                    "error": "Invalid regulation code",
                })
                continue

            semester = Semester.objects.filter(
                degree=department.degree
            ).order_by("sem_number").first()
            if not semester:
                errors.append({
                    "row": row_no,
                    "error": "Semester not found",
                })
                continue

            # -------------------------------------------------
            # 6️⃣ PER-ROW ATOMIC SAVE
            # -------------------------------------------------
            try:
                with transaction.atomic():

                    if User.objects.filter(username=roll_no).exists():
                        errors.append({
                            "row": row_no,
                            "error": "Auth user already exists",
                        })
                        continue

                    user = User.objects.create(
                        username=roll_no,
                        email=student_email,
                        role="STUDENT",
                        is_active=True,
                    )
                    user.set_password(roll_no)
                    user.save()

                    Student.objects.create(
                        user=user,
                        roll_no=roll_no,
                        student_name=student_name,
                        student_email=student_email,
                        student_gender=student_gender,
                        student_date_of_birth=student_date_of_birth,
                        student_phone_number=str(student_phone_number),
                        parent_name=parent_name,
                        parent_phone_number=str(parent_phone_number),
                        batch=regulation.batch,
                        degree=department.degree,
                        department=department,
                        regulation=regulation,
                        semester=semester,
                        section=section,
                        is_active=True,
                    )

                created += 1

            except Exception as e:
                errors.append({
                    "row": row_no,
                    "error": str(e),
                })

        return {
            "created": created,
            "skipped": skipped,
            "errors": errors,
        }


class StudentCreateSerializer(serializers.ModelSerializer):
    dept_code = serializers.CharField(write_only=True)
    regulation_code = serializers.CharField(write_only=True)
    section = serializers.CharField(required=False, allow_null=True, allow_blank=True)

    class Meta:
        model = Student
        fields = [
            "roll_no", "student_name", "student_email", "student_gender",
            "student_date_of_birth", "student_phone_number", "parent_name",
            "parent_phone_number", "dept_code", "regulation_code", "section"
        ]

    @transaction.atomic
    def create(self, validated_data):
        dept_code = validated_data.pop("dept_code")
        regulation_code = validated_data.pop("regulation_code")
        roll_no = validated_data["roll_no"]
        email = validated_data["student_email"]

        # 1. Foreign Key Validations
        department = Department.objects.filter(dept_code__iexact=dept_code).first()
        if not department:
            raise serializers.ValidationError({"dept_code": "Invalid department code"})

        regulation = Regulation.objects.filter(regulation_code__iexact=regulation_code).first()
        if not regulation:
            raise serializers.ValidationError({"regulation_code": "Invalid regulation code"})

        semester = Semester.objects.filter(degree=department.degree).order_by("sem_number").first()
        if not semester:
            raise serializers.ValidationError({"semester": "Initial semester not found for this degree"})

        # 2. User Creation
        if User.objects.filter(username=roll_no).exists():
            raise serializers.ValidationError({"roll_no": "User with this roll number already exists"})

        user = User.objects.create(
            username=roll_no,
            email=email,
            role="STUDENT",
            is_active=True,
        )
        user.set_password(roll_no)
        user.save()

        # 3. Student Creation
        student = Student.objects.create(
            user=user,
            batch=regulation.batch,
            degree=department.degree,
            department=department,
            regulation=regulation,
            semester=semester,
            **validated_data
        )
        return student


class StudentPatchSerializer(serializers.ModelSerializer):
    class Meta:
        model = Student
        fields = [
            "student_name", "student_email", "student_gender", 
            "student_date_of_birth", "student_phone_number", 
            "parent_name", "parent_phone_number", "batch", "section"
        ]

    def update(self, instance, validated_data):
        for field, value in validated_data.items():
            setattr(instance, field, value)
        instance.save()
        return instance



# ==================================================================================
# DEPARTMENT ADMIN ASSIGNMENT SERIALIZER
# ==================================================================================
# This serializer handles the assignment of faculty members as Department Admins.
#
# Key Responsibilities:
# 1. Validate cascading hierarchy (School -> Degree -> Department)
# 2. Ensure selected department actually belongs to selected degree
# 3. Ensure selected degree actually belongs to selected school
# 4. Prevent duplicate assignments (handled by model unique_together)
# 5. Automatically set assigned_by to the requesting user
#
# Cascading Validation Flow:
# - Frontend sends: school_id, degree_id, department_id, faculty_id
# - Backend validates:
#   a) Does the degree belong to the selected school?
#   b) Does the department belong to the selected degree?
# - If valid, create assignment
# - Model's save() method automatically updates faculty user's role
# ==================================================================================

class DepartmentAdminAssignmentSerializer(serializers.ModelSerializer):
    """
    Serializer for creating and managing Department Admin assignments.
    
    Enforces the academic hierarchy: School -> Degree -> Department
    """
    
    # Use PrimaryKeyRelatedField for cleaner input/output
    # Frontend sends UUIDs, backend validates they exist
    faculty_id = serializers.PrimaryKeyRelatedField(
        queryset=Faculty.objects.filter(is_active=True),
        source='faculty',
        help_text="Faculty member to assign (must be active)"
    )
    
    school_id = serializers.PrimaryKeyRelatedField(
        queryset=School.objects.all(),
        source='school',
        help_text="School in the hierarchy"
    )
    
    degree_id = serializers.PrimaryKeyRelatedField(
        queryset=Degree.objects.all(),  # Will be filtered in validation
        source='degree',
        help_text="Degree under the selected school"
    )
    
    department_id = serializers.PrimaryKeyRelatedField(
        queryset=Department.objects.all(),
        source='department',
        help_text="Department under the selected degree"
    )
    
    mfa_id = serializers.UUIDField(
        write_only=True,
        required=True,
        help_text="Verified MFA session ID for this action"
    )
    
    # Read-only fields for output
    assigned_by_email = serializers.EmailField(
        source='assigned_by.email',
        read_only=True,
        help_text="Email of admin who made this assignment"
    )
    
    faculty_name = serializers.CharField(
        source='faculty.full_name',
        read_only=True,
        help_text="Full name of assigned faculty"
    )

    class Meta:
        model = DepartmentAdminAssignment
        fields = [
            'assignment_id', 'faculty_id', 'faculty_name', 'school_id', 
            'degree_id', 'department_id', 'mfa_id', 'assigned_by', 'assigned_by_email',
            'assigned_at', 'is_active'
        ]
        read_only_fields = ['assignment_id', 'assigned_by', 'assigned_at']

    def validate(self, data):
        """
        Validate the cascading hierarchy: School -> Degree -> Department
        
        This ensures:
        1. The selected degree belongs to the selected school
        2. The selected department belongs to the selected degree
        
        If either validation fails, raise a clear error message.
        """
        school = data.get('school')
        degree = data.get('degree')
        department = data.get('department')
        
        # Validation 1: Does the degree belong to the selected school?
        if degree and school:
            if degree.school != school:
                raise serializers.ValidationError({
                    'degree_id': f"Selected degree '{degree.degree_name}' does not belong to school '{school.school_name}'"
                })
        
        # Validation 2: Does the department belong to the selected degree?
        if department and degree:
            if department.degree != degree:
                raise serializers.ValidationError({
                    'department_id': f"Selected department '{department.dept_name}' does not belong to degree '{degree.degree_name}'"
                })
        
        # Validation 3: MFA Verification
        mfa_id = data.get('mfa_id')
        user = self.context['request'].user
        
        try:
            mfa_session = MFASession.objects.get(id=mfa_id, user=user)
        except MFASession.DoesNotExist:
            raise serializers.ValidationError({'mfa_id': "MFA session not found or does not belong to you."})
            
        if not mfa_session.is_verified:
            raise serializers.ValidationError({'mfa_id': "MFA session is not verified. Please approve the push notification."})
            
        if mfa_session.action != 'Department Admin Assignment':
            raise serializers.ValidationError({'mfa_id': f"MFA session was initiated for a different action: {mfa_session.action}"})
            
        # Check expiry (10 mins)
        if timezone.now() > mfa_session.created_at + timezone.timedelta(minutes=10):
            raise serializers.ValidationError({'mfa_id': "MFA session has expired. Please initiate a new one."})
            
        return data

    @transaction.atomic
    def create(self, validated_data):
        """
        Create a new Department Admin assignment.
        
        The model's save() method will automatically update the faculty's
        user role to DEPARTMENT_ADMIN.
        """
        # The assigned_by field should be set by the view from request.user
        # If not already set, we'll handle it in the view
        validated_data.pop('mfa_id', None)
        return DepartmentAdminAssignment.objects.create(**validated_data)
    


"""
--------------------------------------------------------------------------------------------------------------------------------
                                            Roles
--------------------------------------------------------------------------------------------------------------------------------
"""
    
User = get_user_model()

class UserRoleSerializer(serializers.ModelSerializer):
    """
    Serializer for the Roles Dashboard list.
    Unified view for all user roles with their associated profiles.
    """
    profile_details = serializers.SerializerMethodField(read_only=True)
    role_display = serializers.CharField(source='get_role_display', read_only=True)
    all_roles = serializers.SerializerMethodField(read_only=True)

    class Meta:
        model = User
        fields = ['id', 'username', 'email', 'role', 'all_roles', 'role_display', 'profile_details']

    def get_all_roles(self, obj):
        """
        Returns a list of all roles applicable to the user.
        Uses the model's get_all_roles() for consistency.
        """
        return obj.get_all_roles()

    def get_profile_details(self, obj):
        # Faculty or Academic Coordinator with Faculty Profile
        if obj.role in ['FACULTY', 'ACADEMIC_COORDINATOR', 'accedemic_coordinator']:
            faculty = getattr(obj, 'faculty_profile', None)
            if faculty:
                return {
                    'name': faculty.faculty_name,
                    'id': faculty.employee_id,
                    'details': f"Faculty ({faculty.faculty_gender})" + (" (Academic Coordinator)" if obj.role in ["ACADEMIC_COORDINATOR", "accedemic_coordinator"] else "")
                }
        
        # Student Profile
        if obj.role == 'STUDENT':
            student = getattr(obj, 'student_profile', None)
            if student:
                return {
                    'name': student.student_name,
                    'id': student.roll_no,
                    'details': f"Batch: {student.batch}, Dept: {student.department.dept_code if student.department else 'N/A'}",
                    # Academic hierarchy for filtering support in frontend
                    'school_id': str(student.degree.school.school_id) if student.degree and student.degree.school else None,
                    'degree_id': str(student.degree.degree_id) if student.degree else None,
                    'department_id': str(student.department.dept_id) if student.department else None,
                    'batch_name': student.batch
                }
        
        return {
            'name': obj.get_full_name() or obj.username,
            'id': obj.username,
            'details': 'Administrator'
        }
